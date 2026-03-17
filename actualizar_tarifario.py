import pandas as pd
import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def actualizar_hoja(source_file, dest_file_copy, sheet_name, source_skiprows, dest_skiprows, columns_to_rename):
    """
    Actualiza una hoja en el archivo de destino con datos del archivo de origen,
    conservando las macros del archivo original.
    """
    print(f"--- Iniciando actualización de la hoja: {sheet_name} ---")

    try:
        # 1. Leer datos de origen
        print(f"Leyendo hoja '{sheet_name}' de {source_file}...")
        source_df = pd.read_excel(source_file, sheet_name=sheet_name, engine='openpyxl', skiprows=source_skiprows)

        if columns_to_rename:
            source_df.rename(columns=columns_to_rename, inplace=True)

        # 2. Cargar libro de destino y leer datos
        print(f"Cargando libro de trabajo de destino: {dest_file_copy}...")
        book = load_workbook(dest_file_copy, keep_vba=True)
        dest_df = pd.read_excel(dest_file_copy, sheet_name=sheet_name, engine='openpyxl', skiprows=dest_skiprows)
        original_dest_columns = dest_df.columns.tolist()
        dest_df['__original_order__'] = range(len(dest_df))
        
        key_cols = ['CARRIER', 'TRANSPORT ZONE']
        if not all(col in source_df.columns for col in key_cols):
            print(f"Error: Faltan columnas clave en '{source_file}'. Se necesitan: {key_cols}")
            return False
        if not all(col in dest_df.columns for col in key_cols):
            print(f"Error: Faltan columnas clave en el destino. Se necesitan: {key_cols}")
            return False

        # --- Normalización y limpieza ---
        for col in key_cols:
            source_df[col] = source_df[col].astype(str).str.strip()
            dest_df[col] = dest_df[col].astype(str).str.strip()

        source_df.drop_duplicates(subset=key_cols, inplace=True, keep='first')
        dest_df.drop_duplicates(subset=key_cols, inplace=True, keep='first')

        # --- DEBUG: Verificación de claves coincidentes ---
        # Copiamos los dataframes para no alterar los originales con el merge
        temp_source_df = source_df.copy()
        temp_dest_df = dest_df.copy()
        
        # El merge nos dirá cuántas filas coinciden
        merged_df = pd.merge(temp_source_df, temp_dest_df, on=key_cols, how='inner', suffixes=('_origen', '_destino'))
        
        print(f"\n--- DEBUG: Análisis de Coincidencias para '{sheet_name}' ---")
        print(f"Columnas en Origen: {source_df.columns.tolist()}")
        print(f"Columnas en Destino: {dest_df.columns.tolist()}")
        print(f"Filas en Origen (después de limpiar duplicados): {len(source_df)}")
        print(f"Filas en Destino (después de limpiar duplicados): {len(dest_df)}")
        print(f"Se encontraron {len(merged_df)} filas con claves coincidentes ('CARRIER', 'TRANSPORT ZONE') entre el origen y el destino.")
        
        if len(merged_df) > 0:
            print("Ejemplo de claves coincidentes:")
            print(merged_df[key_cols].head())
        else:
            print("¡ATENCIÓN! No se encontraron claves coincidentes. La actualización no modificará ningún dato en esta hoja.")
        print("--- FIN DEBUG ---\n")
        # --- Fin DEBUG ---

        # --- Actualización en memoria ---
        print("Actualizando datos en memoria...")
        dest_df.set_index(key_cols, inplace=True)
        source_df.set_index(key_cols, inplace=True)
        dest_df.update(source_df)
        dest_df.reset_index(inplace=True)

        # --- Reordenar para que coincida con el archivo original ---
        print("Restaurando el orden original de las filas y columnas...")
        dest_df.sort_values('__original_order__', inplace=True)
        dest_df = dest_df[original_dest_columns]

        # 4. Escribir el DataFrame actualizado en la hoja
        print(f"Escribiendo datos actualizados en la hoja '{sheet_name}'...")
        ws = book[sheet_name]
        
        # Borrar solo las filas de datos, manteniendo los encabezados intactos
        # Se asume que los datos comienzan en la fila dest_skiprows + 2
        if ws.max_row > dest_skiprows + 1:
            ws.delete_rows(dest_skiprows + 2, ws.max_row - (dest_skiprows + 1))

        # Escribir filas de datos desde el DataFrame
        for row in dataframe_to_rows(dest_df, index=False, header=False):
            ws.append(row)

        # 5. Guardar el libro
        print(f"Guardando cambios en '{dest_file_copy}'...")
        book.save(dest_file_copy)
        
        print(f"--- Actualización de la hoja '{sheet_name}' completada. ---\n")
        return True

    except (FileNotFoundError, KeyError) as e:
        print(f"Error: No se encontró el archivo o la hoja '{sheet_name}': {e}")
        return False
    except Exception as e:
        print(f"Ocurrió un error inesperado al procesar la hoja '{sheet_name}': {e}")
        return False

# --- Parámetros ---
ARCHIVO_ORIGEN = '01 Tarifario_macro.xlsm'
ARCHIVO_DESTINO_ORIGINAL = '02 Tarifario_macro.xlsm'
ARCHIVO_DESTINO_COPIA = '02_Tarifario_actualizado.xlsm'

CONFIG_HOJAS = [
    {
        'sheet_name': 'B_Tarifas',
        'source_skiprows': 1,
        'dest_skiprows': 1,
        'renames': {'1s': 'Aforo x 900KG'}
    },
    {
        'sheet_name': 'Aumentos',
        'source_skiprows': 11,
        'dest_skiprows': 11,
        'renames': {}
    }
]

# --- Script Principal ---
if __name__ == "__main__":
    print("--- Iniciando script de actualización de tarifarios ---")

    if not os.path.exists(ARCHIVO_DESTINO_ORIGINAL):
        print(f"Error: El archivo de destino original no existe: {ARCHIVO_DESTINO_ORIGINAL}")
    else:
        print(f"Creando una copia de trabajo: {ARCHIVO_DESTINO_COPIA}")
        shutil.copy(ARCHIVO_DESTINO_ORIGINAL, ARCHIVO_DESTINO_COPIA)

        for params in CONFIG_HOJAS:
            success = actualizar_hoja(
                source_file=ARCHIVO_ORIGEN,
                dest_file_copy=ARCHIVO_DESTINO_COPIA,
                sheet_name=params['sheet_name'],
                source_skiprows=params['source_skiprows'],
                dest_skiprows=params['dest_skiprows'],
                columns_to_rename=params['renames']
            )
            if not success:
                print(f"No se pudo procesar la hoja {params['sheet_name']}. Abortando.")
                break
        
        if success:
            print("--- Proceso de actualización completado. ---")
            print(f"El archivo actualizado se ha guardado como: '{ARCHIVO_DESTINO_COPIA}'")